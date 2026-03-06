import os, zipfile, shutil, uuid, re, threading
import cv2, pytesseract, fitz
import numpy as np
import pandas as pd
from flask import Flask, request, render_template_string, send_file, jsonify
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)
UPLOAD_FOLDER = "/tmp/uploads"
REPORT_FOLDER = "/tmp/reports"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

ALLOWED_BILL_EXT = {".jpg", ".jpeg", ".png", ".pdf"}
IGNORE_FOLDERS = {"payment screenshots","payments","misc","receipts","__macosx",".ds_store","payment"}
jobs = {}

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>AuditLens</title>
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;400;500;600&family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet"/>
<style>
:root{--ink:#05080f;--border:rgba(255,255,255,0.07);--gold:#c9a84c;--gold2:#e8cc7a;
  --goldglow:rgba(201,168,76,0.18);--silver:#8892a4;--text:#e8eaf0;--muted:#5a6278;
  --green:#2ecc8f;--red:#e05555;--r:18px;}
*{margin:0;padding:0;box-sizing:border-box}body{font-family:'DM Sans',sans-serif;background:var(--ink);color:var(--text);min-height:100vh;overflow-x:hidden;}
body::before{content:'';position:fixed;inset:0;z-index:0;pointer-events:none;background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.04'/%3E%3C/svg%3E");background-size:200px;opacity:.4;}
.orb{position:fixed;border-radius:50%;filter:blur(120px);pointer-events:none;z-index:0}
.orb1{width:700px;height:700px;background:radial-gradient(circle,rgba(201,168,76,0.06),transparent 70%);top:-200px;right:-100px}
.orb2{width:500px;height:500px;background:radial-gradient(circle,rgba(46,204,143,0.04),transparent 70%);bottom:-100px;left:-100px}
.top-line{position:fixed;top:0;left:0;right:0;height:1px;z-index:100;background:linear-gradient(90deg,transparent,var(--gold) 30%,var(--gold2) 50%,var(--gold) 70%,transparent);}
nav{position:fixed;top:0;left:0;right:0;z-index:50;display:flex;align-items:center;justify-content:space-between;padding:22px 48px;background:rgba(5,8,15,0.7);backdrop-filter:blur(24px);border-bottom:1px solid var(--border);}
.nav-brand{display:flex;align-items:center;gap:14px;}.nav-logo-box{width:36px;height:36px;border-radius:8px;background:linear-gradient(135deg,var(--gold),#8b6914);display:flex;align-items:center;justify-content:center;font-size:16px;box-shadow:0 0 20px var(--goldglow);}
.nav-name{font-family:'Cormorant Garamond',serif;font-size:22px;font-weight:500;color:#fff;}.nav-name span{color:var(--gold)}
.nav-pills{display:flex;gap:8px;}.nav-pill{font-size:11px;letter-spacing:1.5px;text-transform:uppercase;padding:5px 14px;border-radius:999px;color:var(--silver);border:1px solid var(--border);background:rgba(255,255,255,0.02);}
.nav-pill.live{color:var(--green);border-color:rgba(46,204,143,0.3);background:rgba(46,204,143,0.06);}
.nav-pill.live::before{content:'';display:inline-block;width:6px;height:6px;border-radius:50%;background:var(--green);margin-right:7px;animation:pulse-dot 2s ease infinite;}
@keyframes pulse-dot{0%,100%{opacity:1;box-shadow:0 0 0 0 rgba(46,204,143,.4)}50%{opacity:.7;box-shadow:0 0 0 4px rgba(46,204,143,0)}}
.page{position:relative;z-index:1;max-width:860px;margin:0 auto;padding:140px 24px 100px;}
.hero{text-align:center;margin-bottom:72px}.hero-eyebrow{display:inline-flex;align-items:center;gap:10px;font-size:11px;letter-spacing:3px;text-transform:uppercase;color:var(--gold);margin-bottom:28px;}
.hero-eyebrow::before,.hero-eyebrow::after{content:'';flex:1;height:1px;width:40px;background:linear-gradient(90deg,transparent,var(--gold));}.hero-eyebrow::after{transform:scaleX(-1)}
.hero h1{font-family:'Cormorant Garamond',serif;font-size:clamp(48px,7vw,80px);font-weight:300;line-height:1.05;letter-spacing:-1px;color:#fff;margin-bottom:24px;}
.hero h1 em{font-style:italic;background:linear-gradient(135deg,var(--gold2),var(--gold));-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.hero-sub{font-size:16px;color:var(--silver);line-height:1.8;max-width:480px;margin:0 auto;font-weight:300;}
.panel{background:rgba(255,255,255,0.025);border:1px solid var(--border);border-radius:var(--r);backdrop-filter:blur(20px);position:relative;overflow:hidden;margin-bottom:20px;}
.panel::before{content:'';position:absolute;top:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,rgba(201,168,76,0.4),transparent);}
.panel-body{padding:36px 40px}.section-label{display:flex;align-items:center;gap:12px;font-size:11px;letter-spacing:2.5px;text-transform:uppercase;color:var(--gold);font-weight:500;margin-bottom:28px;}
.section-label::after{content:'';flex:1;height:1px;background:var(--border)}
.drop-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}
.drop-zone{border:1px dashed rgba(201,168,76,0.2);border-radius:12px;padding:32px 20px;text-align:center;cursor:pointer;transition:all .3s;background:rgba(201,168,76,0.02);position:relative;overflow:hidden;}
.drop-zone:hover{border-color:rgba(201,168,76,0.5);background:rgba(201,168,76,0.04);transform:translateY(-2px);}
.drop-zone.has-file{border-color:rgba(46,204,143,0.4);background:rgba(46,204,143,0.03);}
.dz-icon{font-size:32px;margin-bottom:12px;display:block}.dz-title{font-size:13px;font-weight:500;color:var(--text);margin-bottom:5px}.dz-sub{font-size:11px;color:var(--muted)}
.dz-file{margin-top:10px;font-size:12px;color:var(--green);font-family:'DM Mono',monospace;display:none;}.dz-file.visible{display:block}
input[type="file"]{display:none}
.run-btn{width:100%;padding:18px;border:none;border-radius:12px;background:linear-gradient(135deg,#c9a84c,#e8cc7a 40%,#c9a84c);color:#0a0804;font-family:'DM Sans',sans-serif;font-size:15px;font-weight:600;cursor:pointer;transition:all .3s;position:relative;overflow:hidden;}
.run-btn::before{content:'';position:absolute;top:0;left:-100%;width:100%;height:100%;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.2),transparent);transition:left .5s;}
.run-btn:hover:not(:disabled)::before{left:100%}.run-btn:hover:not(:disabled){transform:translateY(-2px);box-shadow:0 8px 30px rgba(201,168,76,0.4);}
.run-btn:disabled{background:rgba(255,255,255,0.05);color:var(--muted);cursor:not-allowed;}
.err-box{background:rgba(224,85,85,0.08);border:1px solid rgba(224,85,85,0.25);border-radius:10px;padding:14px 18px;color:#ff8080;display:none;margin-top:16px;font-size:13px;font-family:'DM Mono',monospace;}
#progressPanel{display:none}.prog-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:16px}
.prog-label{font-size:13px;color:var(--gold);font-weight:500}.prog-pct{font-family:'Cormorant Garamond',serif;font-size:28px;font-weight:300;color:#fff;}
.prog-track{height:2px;background:rgba(255,255,255,0.06);border-radius:999px;overflow:hidden;margin-bottom:12px;}
.prog-fill{height:100%;border-radius:999px;width:0%;background:linear-gradient(90deg,var(--gold),var(--gold2));transition:width .7s;position:relative;}
.prog-fill::after{content:'';position:absolute;right:0;top:-3px;width:8px;height:8px;border-radius:50%;background:var(--gold2);box-shadow:0 0 10px var(--gold);}
.prog-step{font-size:12px;color:var(--muted);font-family:'DM Mono',monospace;text-align:right;min-height:18px;margin-bottom:20px;}
.terminal{background:#020408;border:1px solid rgba(255,255,255,0.06);border-radius:10px;overflow:hidden;}
.terminal-bar{background:rgba(255,255,255,0.03);border-bottom:1px solid rgba(255,255,255,0.05);padding:10px 16px;display:flex;align-items:center;gap:8px;}
.t-dot{width:10px;height:10px;border-radius:50%}.t-dot:nth-child(1){background:#e05555}.t-dot:nth-child(2){background:#e0b455}.t-dot:nth-child(3){background:#2ecc8f}
.terminal-title{font-size:11px;color:var(--muted);letter-spacing:1px;text-transform:uppercase;margin-left:8px;font-family:'DM Mono',monospace;}
.terminal-body{padding:16px;max-height:260px;overflow-y:auto;font-family:'DM Mono',monospace;font-size:12px;line-height:1.9;}
.terminal-body::-webkit-scrollbar{width:3px}.terminal-body::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px}
.log-line{padding:1px 0}.log-line.ok{color:#2ecc8f}.log-line.err{color:#e05555}.log-line.info{color:#5a6278}
.log-line.ok::before{content:'+ '}.log-line.err::before{content:'! '}.log-line.info::before{content:'  '}
.cursor-blink{display:inline-block;width:7px;height:14px;background:var(--gold);vertical-align:middle;margin-left:2px;animation:blink 1s step-end infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:0}}
#resultPanel{display:none}.metrics-row{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:32px;}
.metric{background:rgba(255,255,255,0.02);border:1px solid var(--border);border-radius:14px;padding:24px 20px;text-align:center;transition:transform .2s;position:relative;overflow:hidden;}
.metric::before{content:'';position:absolute;bottom:0;left:0;right:0;height:1px;}
.metric.m-total::before{background:linear-gradient(90deg,transparent,rgba(136,146,255,.5),transparent)}
.metric.m-match::before{background:linear-gradient(90deg,transparent,rgba(46,204,143,.5),transparent)}
.metric.m-miss::before{background:linear-gradient(90deg,transparent,rgba(224,85,85,.5),transparent)}
.metric:hover{transform:translateY(-3px)}.metric-num{font-family:'Cormorant Garamond',serif;font-size:52px;font-weight:300;line-height:1;margin-bottom:8px;letter-spacing:-2px;}
.metric.m-total .metric-num{color:#8892ff}.metric.m-match .metric-num{color:var(--green)}.metric.m-miss .metric-num{color:var(--red)}
.metric-label{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:var(--muted);font-weight:500;}
.dl-btn{width:100%;padding:18px;border:none;border-radius:12px;background:transparent;border:1px solid rgba(46,204,143,0.3);color:var(--green);font-family:'DM Sans',sans-serif;font-size:15px;cursor:pointer;transition:all .3s;display:flex;align-items:center;justify-content:center;gap:12px;}
.dl-btn:hover{background:rgba(46,204,143,0.08);border-color:rgba(46,204,143,0.6);box-shadow:0 0 30px rgba(46,204,143,.1);transform:translateY(-2px);}
.steps-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}
.step-card{background:rgba(255,255,255,0.02);border:1px solid var(--border);border-radius:12px;padding:22px 18px;transition:all .3s;}
.step-card:hover{border-color:rgba(201,168,76,0.25);transform:translateY(-3px);}
.step-num{font-family:'Cormorant Garamond',serif;font-size:36px;font-weight:300;color:rgba(201,168,76,0.25);line-height:1;margin-bottom:12px;}
.step-title{font-size:13px;font-weight:600;color:var(--text);margin-bottom:6px}.step-desc{font-size:12px;color:var(--muted);line-height:1.7}
footer{text-align:center;padding:40px 0 60px;position:relative;z-index:1;}
.footer-brand{font-family:'Cormorant Garamond',serif;font-size:13px;color:var(--muted);}.footer-brand strong{color:var(--gold);font-weight:400}
.footer-rule{width:60px;height:1px;margin:16px auto;background:linear-gradient(90deg,transparent,var(--border),transparent);}
@keyframes fadeUp{from{opacity:0;transform:translateY(24px)}to{opacity:1;transform:translateY(0)}}
.fade-up{animation:fadeUp .7s cubic-bezier(.4,0,.2,1) both}
@media(max-width:640px){.drop-row,.steps-grid{grid-template-columns:1fr 1fr}nav{padding:18px 20px}.page{padding:120px 16px 80px}.panel-body{padding:24px 20px}.metrics-row{grid-template-columns:1fr}}
</style></head><body>
<div class="top-line"></div><div class="orb orb1"></div><div class="orb orb2"></div>
<nav>
  <div class="nav-brand"><div class="nav-logo-box">&#x1F9FE;</div><div class="nav-name">Audit<span>Lens</span></div></div>
  <div class="nav-pills"><div class="nav-pill">v4.0</div><div class="nav-pill live">System Online</div></div>
</nav>
<div class="page">
  <div class="hero fade-up">
    <div class="hero-eyebrow">Purchase Intelligence Platform</div>
    <h1>Audit Every Bill.<br><em>Instantly.</em></h1>
    <p class="hero-sub">Upload your bill images and Zoho export. OCR scans every document, matches records with precision, and delivers a verified audit report.</p>
  </div>
  <div class="panel fade-up"><div class="panel-body">
    <div class="section-label">Upload Files</div>
    <div class="drop-row">
      <div class="drop-zone" id="zipZone" onclick="document.getElementById('fileInput').click()">
        <div class="dz-icon">&#x1F5C2;</div><div class="dz-title">Bill Images ZIP</div><div class="dz-sub">JPG &middot; PNG &middot; PDF</div>
        <div class="dz-file" id="zipLabel"></div><input type="file" id="fileInput" accept=".zip" onchange="handleZip(this)"/>
      </div>
      <div class="drop-zone" id="csvZone" onclick="document.getElementById('csvInput').click()">
        <div class="dz-icon">&#x1F4CA;</div><div class="dz-title">Zoho Excel / CSV</div><div class="dz-sub">Reference for matching</div>
        <div class="dz-file" id="csvLabel"></div><input type="file" id="csvInput" accept=".csv,.xlsx,.xls" onchange="handleCsv(this)"/>
      </div>
    </div>
    <button class="run-btn" id="runBtn" onclick="startAudit()" disabled>Begin Audit</button>
    <div class="err-box" id="errBox"></div>
  </div></div>
  <div class="panel" id="progressPanel"><div class="panel-body">
    <div class="section-label">Processing</div>
    <div class="prog-header"><span class="prog-label" id="progLabel">Initialising...</span><span class="prog-pct"><span id="pctNum">0</span>%</span></div>
    <div class="prog-track"><div class="prog-fill" id="progFill"></div></div>
    <div class="prog-step" id="progStep">&nbsp;</div>
    <div class="terminal">
      <div class="terminal-bar"><div class="t-dot"></div><div class="t-dot"></div><div class="t-dot"></div><span class="terminal-title">OCR Output Stream</span></div>
      <div class="terminal-body" id="logArea"><span class="cursor-blink"></span></div>
    </div>
  </div></div>
  <div class="panel" id="resultPanel"><div class="panel-body">
    <div class="section-label">Audit Results</div>
    <div class="metrics-row">
      <div class="metric m-total"><div class="metric-num" id="statTotal">0</div><div class="metric-label">Records Audited</div></div>
      <div class="metric m-match"><div class="metric-num" id="statMatched">0</div><div class="metric-label">Confirmed Matched</div></div>
      <div class="metric m-miss"><div class="metric-num" id="statMismatch">0</div><div class="metric-label">Requires Review</div></div>
    </div>
    <button class="dl-btn" id="dlBtn"><span>&#x2913;</span> Download Excel Report</button>
  </div></div>
  <div class="panel fade-up"><div class="panel-body">
    <div class="section-label">How It Works</div>
    <div class="steps-grid">
      <div class="step-card"><div class="step-num">01</div><div class="step-title">Upload</div><div class="step-desc">Drop your ZIP of bill images and your Zoho CSV export.</div></div>
      <div class="step-card"><div class="step-num">02</div><div class="step-title">OCR Scan</div><div class="step-desc">2-pass OCR extracts text from every bill — digital or photographed.</div></div>
      <div class="step-card"><div class="step-num">03</div><div class="step-title">Smart Match</div><div class="step-desc">Bill number, vendor, item and amount matched with fuzzy intelligence.</div></div>
      <div class="step-card"><div class="step-num">04</div><div class="step-title">Report</div><div class="step-desc">Colour-coded Excel: green matched, yellow review, red not found.</div></div>
    </div>
  </div></div>
</div>
<footer><div class="footer-rule"></div><div class="footer-brand">AuditLens &mdash; Built by <strong>Shubham Hulsure</strong></div></footer>
<script>
let zipFile=null,currentJobId=null,pollTimer=null;
function handleZip(i){zipFile=i.files[0];const l=document.getElementById('zipLabel'),z=document.getElementById('zipZone');if(zipFile){l.textContent=zipFile.name;l.classList.add('visible');z.classList.add('has-file');}checkReady();}
function handleCsv(i){const l=document.getElementById('csvLabel'),z=document.getElementById('csvZone');if(i.files[0]){l.textContent=i.files[0].name;l.classList.add('visible');z.classList.add('has-file');}checkReady();}
function checkReady(){document.getElementById('runBtn').disabled=!zipFile;}
const zz=document.getElementById('zipZone');
zz.addEventListener('dragover',e=>{e.preventDefault();zz.classList.add('dragover');});
zz.addEventListener('dragleave',()=>zz.classList.remove('dragover'));
zz.addEventListener('drop',e=>{e.preventDefault();zz.classList.remove('dragover');const f=e.dataTransfer.files[0];if(f&&f.name.endsWith('.zip')){zipFile=f;document.getElementById('zipLabel').textContent=f.name;document.getElementById('zipLabel').classList.add('visible');zz.classList.add('has-file');checkReady();}});
function addLog(msg,type){const log=document.getElementById('logArea');const cur=log.querySelector('.cursor-blink');if(cur)cur.remove();const d=document.createElement('div');d.className='log-line '+(type||'info');d.textContent=msg;log.appendChild(d);const c=document.createElement('span');c.className='cursor-blink';log.appendChild(c);log.scrollTop=log.scrollHeight;}
function showErr(m){const b=document.getElementById('errBox');b.textContent=m;b.style.display='block';}
function animateNum(el,target){const dur=1400,t0=performance.now();function step(now){const p=Math.min((now-t0)/dur,1),e=1-Math.pow(1-p,3);el.textContent=Math.floor(e*target);if(p<1)requestAnimationFrame(step);else el.textContent=target;}requestAnimationFrame(step);}
async function startAudit(){if(!zipFile)return;document.getElementById('errBox').style.display='none';document.getElementById('logArea').innerHTML='<span class="cursor-blink"></span>';document.getElementById('resultPanel').style.display='none';const btn=document.getElementById('runBtn');btn.disabled=true;btn.textContent='Uploading...';document.getElementById('progressPanel').style.display='block';document.getElementById('progFill').style.width='3%';document.getElementById('pctNum').textContent='3';document.getElementById('progLabel').textContent='Uploading...';const fd=new FormData();fd.append('zip_file',zipFile);const cf=document.getElementById('csvInput').files[0];if(cf)fd.append('csv_file',cf);try{const res=await fetch('/start',{method:'POST',body:fd});const data=await res.json();if(!data.job_id){showErr(data.error||'Upload failed');btn.disabled=false;btn.textContent='Begin Audit';return;}currentJobId=data.job_id;btn.textContent='Processing...';addLog('Files uploaded. Starting OCR...','ok');pollTimer=setInterval(pollStatus,2500);}catch(e){showErr(e.message);btn.disabled=false;btn.textContent='Begin Audit';}}
async function pollStatus(){if(!currentJobId)return;try{const res=await fetch('/status/'+currentJobId);const job=await res.json();const pct=job.progress||0;document.getElementById('progFill').style.width=pct+'%';document.getElementById('pctNum').textContent=pct;document.getElementById('progStep').textContent=job.step||'';document.getElementById('progLabel').textContent=pct<20?'Extracting...':pct<68?'OCR scanning...':pct<90?'Matching records...':'Writing report...';if(job.new_logs&&job.new_logs.length)job.new_logs.forEach(l=>addLog(l.msg,l.type));if(job.status==='done'){clearInterval(pollTimer);const r=document.getElementById('resultPanel');r.style.display='block';r.scrollIntoView({behavior:'smooth',block:'start'});animateNum(document.getElementById('statTotal'),job.summary.total);animateNum(document.getElementById('statMatched'),job.summary.matched);animateNum(document.getElementById('statMismatch'),job.summary.mismatch);document.getElementById('dlBtn').onclick=()=>{window.location.href='/download/'+currentJobId;};document.getElementById('runBtn').disabled=false;document.getElementById('runBtn').textContent='Begin Audit';}if(job.status==='error'){clearInterval(pollTimer);showErr(job.error||'Something went wrong.');document.getElementById('runBtn').disabled=false;document.getElementById('runBtn').textContent='Begin Audit';}}catch(e){}}
</script></body></html>"""

# ── Helpers ───────────────────────────────────────────────────────────────────

def is_ignored(path):
    parts = [p.strip().lower() for p in path.replace("\\", "/").split("/")]
    return any(p in IGNORE_FOLDERS for p in parts)

def normalize(text):
    if not text: return ""
    text = re.sub(r"[^a-z0-9\s]", " ", str(text).lower())
    return re.sub(r"\s+", " ", text).strip()

def normalize_billno(text):
    return re.sub(r"[^a-z0-9]", "", str(text).lower())

def extract_numbers(text):
    return set(re.findall(r"\b\d+(?:\.\d+)?\b", text))

def pdf_to_images(pdf_path, output_dir):
    images = []
    try:
        doc = fitz.open(pdf_path)
        for i, page in enumerate(doc):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(output_dir, f"pdf_page_{i}.jpg")
            pix.save(img_path)
            images.append(img_path)
    except Exception as e:
        print(f"PDF error: {e}")
    return images

# ── OCR: 2-pass, fast, accurate ───────────────────────────────────────────────

SHARP_KERNEL = np.array([[0,-1,0],[-1,5,-1],[0,-1,0]])

def ocr_image(path):
    """
    2-pass OCR, optimised for speed + accuracy on phone-photographed bills.

    Pass 1: Raw colour image — fast, works perfectly for clean digital/printed bills
    Pass 2: 2x upscale + sharpen + adaptive threshold — catches handwritten/blurry bills

    Both run sequentially (parallel gave no real speed benefit due to tesseract's
    internal threading). Combined text gives both clean digital reads AND noisy
    photo reads — the fuzzy matcher uses whichever signals are present.

    NO denoising (fastNlMeansDenoising was the 3-hour bottleneck — removed entirely).
    """
    try:
        img = cv2.imread(path)
        if img is None:
            return ""
        h, w = img.shape[:2]

        # ── Pass 1: raw colour (best for clean printed bills, ~12s)
        t1 = pytesseract.image_to_string(img, config="--psm 6 --oem 1")

        # ── Pass 2: 2x upscale + sharpen on gray + adaptive threshold
        #    (best for blurry/photographed bills, ~20s)
        #    Only upscale if image is small — saves time on already-large images
        if max(h, w) < 1600:
            big = cv2.resize(img, (w*2, h*2), interpolation=cv2.INTER_CUBIC)
        else:
            big = img
        gray = cv2.cvtColor(big, cv2.COLOR_BGR2GRAY)
        sharp = cv2.filter2D(gray, -1, SHARP_KERNEL)
        ada = cv2.adaptiveThreshold(
            sharp, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 31, 10
        )
        t2 = pytesseract.image_to_string(ada, config="--psm 6 --oem 1")

        # Combine both passes
        combined = normalize(t1) + " " + normalize(t2)
        return combined

    except Exception as e:
        print(f"OCR error {path}: {e}")
        return ""

# ── Scoring ───────────────────────────────────────────────────────────────────

def bill_number_score(bill_no_raw, ocr_text):
    """Try multiple forms of the bill number with word boundaries."""
    if not bill_no_raw or len(bill_no_raw) < 2:
        return 0, False
    forms = {
        normalize(bill_no_raw),          # "cr 482"
        normalize_billno(bill_no_raw),   # "cr482"
    }
    digits = re.sub(r"[^0-9]", "", bill_no_raw)
    if len(digits) >= 3:
        forms.add(digits)               # "482"
    forms.discard("")

    for form in sorted(forms, key=len, reverse=True):  # longest form first
        if re.search(r'\b' + re.escape(form) + r'\b', ocr_text):
            if form.isdigit() and len(form) <= 4:
                return 35, True
            elif form.isdigit():
                return 50, True
            else:
                return 65, True

    best = max((fuzz.partial_ratio(f, ocr_text) for f in forms), default=0)
    return best * 0.20, False


def score_match(bill_no, vendor, item, amount_str, ocr_text):
    """
    Score 0–100.
    Bill no  → 0–65 pts
    Vendor   → 0–20 pts  (token_set_ratio handles word-order + OCR garbling)
    Item     → 0–15 pts  (token_set_ratio handles partial reads)
    Amount   → 0–10 pts
    """
    if not ocr_text:
        return 0
    score, _ = bill_number_score(bill_no, ocr_text)
    if vendor and len(vendor) >= 2:
        # token_set_ratio is more forgiving of garbled/reordered OCR text
        score += fuzz.token_set_ratio(vendor, ocr_text) * 0.20
    if item and len(item) >= 2:
        score += fuzz.token_set_ratio(item, ocr_text) * 0.15
    if amount_str:
        amt = re.sub(r"\.0+$", "", amount_str.strip())
        if amt in extract_numbers(ocr_text):
            score += 10
    return min(round(score, 1), 100)


def classify(score):
    """
    Threshold set at 55 based on analysis of real invoice OCR scores.
    Real matches consistently score 60–75 even with garbled vendor names.
    False positives rarely exceed 45 with multi-field scoring.
    """
    if score >= 55:
        return "Matched",              "Match confirmed"
    elif score >= 35:
        return "Mismatch / Duplicate", "Partial match — verify manually"
    else:
        return "Not Found",            "No matching bill image detected"

# ── Excel writer ──────────────────────────────────────────────────────────────

def generate_excel(results, output_path):
    wb = Workbook(); ws = wb.active; ws.title = "Audit Report"
    headers = ["File Name","Folder","Bill Number","Bill Date","Vendor Name",
               "Customer/Hotel","Item Description","Quantity","Rate (Rs)",
               "Total Amount (Rs)","AI Confidence","Match Status","Match Detail"]
    hfill = PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    hfont = Font(bold=True,color="FFFFFF",size=11)
    for col,h in enumerate(headers,1):
        c=ws.cell(row=1,column=col,value=h); c.fill=hfill; c.font=hfont
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30
    green  = PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
    red    = PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")
    for i,r in enumerate(results,2):
        row_data=[r.get("file_name",""),r.get("folder",""),r.get("bill_number",""),
                  r.get("bill_date",""),r.get("vendor_name",""),r.get("customer_name",""),
                  r.get("item_description",""),r.get("quantity",""),r.get("rate",""),
                  r.get("total_amount",""),r.get("confidence",""),r.get("match_status",""),r.get("match_detail","")]
        s=r.get("match_status","")
        rf=green if s=="Matched" else red if s=="Not Found" else yellow if "Mismatch" in s else None
        for col,val in enumerate(row_data,1):
            c=ws.cell(row=i,column=col,value=val); c.alignment=Alignment(vertical="center")
            if rf: c.fill=rf
    col_widths=[22,18,15,12,30,25,25,10,12,15,12,14,35]
    for col,width in enumerate(col_widths,1):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=width
    ws.freeze_panes="A2"; wb.save(output_path)

# ── Background worker ─────────────────────────────────────────────────────────

def run_audit(job_id, work_dir, zip_path, csv_path):
    job = jobs[job_id]
    def log(msg, t="info"):
        job["all_logs"].append({"msg":msg,"type":t})
        job["new_logs"].append({"msg":msg,"type":t})
    def update(progress, step):
        job["progress"]=progress; job["step"]=step

    try:
        update(8,"Extracting ZIP...")
        extract_dir=os.path.join(work_dir,"extracted")
        os.makedirs(extract_dir,exist_ok=True)
        with zipfile.ZipFile(zip_path,"r") as z: z.extractall(extract_dir)

        csv_df=None
        if csv_path and os.path.exists(csv_path):
            update(12,"Reading reference file...")
            try:
                csv_df=pd.read_csv(csv_path) if csv_path.endswith(".csv") else pd.read_excel(csv_path,engine="openpyxl")
                csv_df.columns=csv_df.columns.str.strip()
                log(f"Reference file: {len(csv_df)} rows","ok")
            except Exception as e:
                log(f"Reference file error: {e}","err")

        update(16,"Scanning for images...")
        bill_images=[]
        for root,dirs,files in os.walk(extract_dir):
            for f in files:
                rel=os.path.relpath(os.path.join(root,f),extract_dir)
                full=os.path.join(root,f)
                if is_ignored(rel): continue
                ext=os.path.splitext(f)[1].lower()
                if ext==".pdf": bill_images.extend([(img,rel) for img in pdf_to_images(full,root)])
                elif ext in ALLOWED_BILL_EXT: bill_images.append((full,rel))

        total_imgs=len(bill_images)
        log(f"Found {total_imgs} bill images — 2-pass OCR starting","info")

        ocr_cache=[]
        for i,(img_path,rel_path) in enumerate(bill_images):
            update(16+int((i/max(total_imgs,1))*50),
                   f"OCR {i+1}/{total_imgs}: {os.path.basename(rel_path)}")
            txt=ocr_image(img_path)
            if txt and len(txt.strip())>10:
                ocr_cache.append((img_path,txt))
                log(f"OCR OK: {os.path.basename(rel_path)} ({len(txt)} chars)","ok")
            else:
                log(f"OCR empty: {os.path.basename(rel_path)}","err")
            job["new_logs"]=job["new_logs"][-30:]

        results=[]
        if csv_df is not None:
            update(68,"Matching records to bills...")
            log("Matching started...","info")
            total_rows=len(csv_df)
            for idx,(_,row) in enumerate(csv_df.iterrows()):
                update(68+int((idx/max(total_rows,1))*22),f"Row {idx+1}/{total_rows}")
                bill_no=str(row.get("Bill Number",""))
                vendor =normalize(str(row.get("Vendor Name","")))
                item   =normalize(str(row.get("Item Name","")))
                amount =normalize(str(row.get("Item Total","")))
                best_score=0; best_path=""; best_signals=""
                for path,text in ocr_cache:
                    s=score_match(bill_no,vendor,item,amount,text)
                    if s>best_score:
                        best_score=s; best_path=path
                        sigs=[]
                        forms={normalize(bill_no),normalize_billno(bill_no)}
                        digs=re.sub(r"[^0-9]","",bill_no)
                        if len(digs)>=3: forms.add(digs)
                        for f in forms:
                            if f and re.search(r'\b'+re.escape(f)+r'\b',text):
                                sigs.append(f"BillNo:{bill_no}"); break
                        vr=fuzz.token_set_ratio(vendor,text)
                        if vr>=60: sigs.append(f"Vendor:{vr}%")
                        amt=re.sub(r"\.0+$","",amount.strip())
                        if amt and amt in extract_numbers(text): sigs.append(f"Amt:{amt}")
                        best_signals=" | ".join(sigs) if sigs else "item fuzzy only"
                status,_=classify(best_score)
                results.append({
                    "file_name":os.path.basename(best_path) if best_path else "",
                    "folder":os.path.dirname(best_path) if best_path else "",
                    "bill_number":str(row.get("Bill Number","")),
                    "bill_date":str(row.get("Bill Date","")),
                    "vendor_name":str(row.get("Vendor Name","")),
                    "customer_name":str(row.get("Branch Name","")),
                    "item_description":str(row.get("Item Name","")),
                    "quantity":str(row.get("Quantity","")),
                    "rate":str(row.get("Rate","")),
                    "total_amount":str(row.get("Item Total","")),
                    "confidence":best_score,
                    "match_status":status,
                    "match_detail":best_signals,
                })
                t="ok" if status=="Matched" else ("err" if status=="Not Found" else "info")
                log(f"{status} ({best_score}) | {bill_no} | {str(row.get('Vendor Name',''))[:15]} | {best_signals}",t)
        else:
            for img_path,_ in ocr_cache:
                results.append({"file_name":os.path.basename(img_path),"folder":"","bill_number":"",
                    "bill_date":"","vendor_name":"","customer_name":"","item_description":"",
                    "quantity":"","rate":"","total_amount":"","confidence":"","match_status":"No CSV","match_detail":""})

        update(92,"Writing report...")
        report_path=os.path.join(REPORT_FOLDER,f"audit_{job_id}.xlsx")
        generate_excel(results,report_path)
        matched =sum(1 for r in results if r["match_status"]=="Matched")
        mismatch=sum(1 for r in results if r["match_status"] in ("Not Found","Mismatch / Duplicate"))
        log(f"Done — {matched} matched, {mismatch} flagged, {len(results)} total","ok")
        job.update({"status":"done","progress":100,"step":"Complete",
                    "report_path":report_path,
                    "summary":{"total":len(results),"matched":matched,"mismatch":mismatch}})
    except Exception as e:
        import traceback
        job.update({"status":"error","error":str(e)+"\n"+traceback.format_exc()})
    finally:
        shutil.rmtree(work_dir,ignore_errors=True)

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index(): return render_template_string(HTML)

@app.route("/start",methods=["POST"])
def start():
    job_id=str(uuid.uuid4())[:8]; work_dir=os.path.join(UPLOAD_FOLDER,job_id)
    os.makedirs(work_dir,exist_ok=True)
    zip_file=request.files.get("zip_file")
    if not zip_file: return jsonify({"error":"No ZIP file"}),400
    zip_path=os.path.join(work_dir,"bills.zip"); zip_file.save(zip_path)
    csv_path=None; csv_file=request.files.get("csv_file")
    if csv_file:
        ext=os.path.splitext(csv_file.filename)[1].lower() or ".xlsx"
        csv_path=os.path.join(work_dir,f"data{ext}"); csv_file.save(csv_path)
    jobs[job_id]={"status":"processing","progress":5,"step":"Files received...",
                  "all_logs":[],"new_logs":[],"summary":{},"report_path":None,"error":None}
    t=threading.Thread(target=run_audit,args=(job_id,work_dir,zip_path,csv_path))
    t.daemon=True; t.start()
    return jsonify({"job_id":job_id})

@app.route("/status/<job_id>")
def status(job_id):
    if job_id not in jobs: return jsonify({"status":"not_found"}),404
    job=jobs[job_id]
    out={k:job[k] for k in ("status","progress","step","summary","error")}
    out["new_logs"]=job["new_logs"]; job["new_logs"]=[]
    return jsonify(out)

@app.route("/download/<job_id>")
def download(job_id):
    if job_id not in jobs or not jobs[job_id].get("report_path"): return "Not ready",404
    rp=jobs[job_id]["report_path"]
    if os.path.exists(rp): return send_file(rp,as_attachment=True,download_name="Purchase_Audit_Report.xlsx")
    return "File not found",404

if __name__=="__main__":
    port=int(os.environ.get("PORT",8000))
    app.run(debug=False,host="0.0.0.0",port=port)
